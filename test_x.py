from openpyxl import load_workbook
from typing import List, Dict, Any, Optional

class ExcelReader:
    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        """
        ExcelReaderを初期化する

        Args:
            file_path (str): 読み取るExcelファイルのパス
            sheet_name (Optional[str]): 使用するシート名指定がない場合はアクティブシートを使用
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self._load_workbook()

    def _load_workbook(self):
        """Excelファイルをロードし、指定されたシートを選択する"""
        try:
            self.workbook = load_workbook(self.file_path, read_only=True)
            self.sheet = self.workbook[self.sheet_name] if self.sheet_name else self.workbook.active
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {self.file_path}")
        except KeyError:
            raise KeyError(f"Sheet not found: {self.sheet_name}")

    def get_headers(self) -> List[str]:
        """
        先頭行をヘッダーとして取得する

        Returns:
            List[str]: ヘッダー名のリスト
        """
        return [cell.value for cell in self.sheet[1]]

    def read_all(self) -> List[Dict[str, Any]]:
        """
        シート内の全データを辞書形式で取得する

        Returns:
            List[Dict[str, Any]]: データのリスト（各行は辞書形式）
        """
        headers = self.get_headers()
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return data

    def find_rows(self, column: str, value: Any) -> List[Dict[str, Any]]:
        """
        指定した列が特定の値と一致する行を検索する

        Args:
            column (str): 検索対象の列名
            value (Any): 一致させたい値

        Returns:
            List[Dict[str, Any]]: 条件に一致する行のリスト
        """
        headers = self.get_headers()
        if column not in headers:
            raise KeyError(f"Column not found: {column}")
        data = self.read_all()
        return [row for row in data if row.get(column) == value]

    def __enter__(self):
        """Context Managerのサポート"""
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        """Context Managerで使用後にリソースを解放する"""
        if self.workbook:
            self.workbook.close()

if __name__ == "__main__":
    # 使用例
    # 1. 全データの読み取り
    reader = ExcelReader('example.xlsx')
    data = reader.read_all()
    print(data)

    # 2. 特定の値を含む行の検索
    reader = ExcelReader('example.xlsx')
    results = reader.find_rows('Name', 'Alice')
    print(results)

    # 3. Context Managerを利用して安全に操作
    with ExcelReader('example.xlsx') as reader:
        headers = reader.get_headers()
        print(f"Headers: {headers}")
        data = reader.read_all()
        
    print(data)